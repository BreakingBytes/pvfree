from django.db import models, IntegrityError
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from datetime import date
from tastypie.resources import ModelResource
from tastypie.authorization import DjangoAuthorization
from tastypie.authentication import ApiKeyAuthentication
from django.contrib.auth.models import User
from django.db.models import signals
from tastypie.models import create_api_key
import logging

LOGGER = logging.getLogger(__name__)

signals.post_save.connect(create_api_key, sender=User)


MAPPING = {
    "Unique ID#": "Sandia_ID",
    "Manufacturer": "manufacturer",
    "ID": "name",
    "Source": "source",
    "Vintage": "vintage",
    "ac Voltage": "Vaco",
    "Paco": "Paco",
    "Pdco": "Pdco",
    "Vdco": "Vdco",
    "Pso": "Pso",
    "Co": "C0",
    "C1": "C1",
    "C2": "C2",
    "C3": "C3",
    "Vdcmax": "Vdcmax",
    "Idcmax": "Idcmax",
    "MPPT-Low": "MPPT_low",
    "MPPT-Hi": "MPPT_hi",
    "Pnt": "Pnt",
    "Tamb Low": "Tamb_low",
    "Tamb Max": "Tamb_max",
    "Weight": "weight",
    }


class ApiKeyAuthOrReadOnly(ApiKeyAuthentication):
    def _unauthorized(self):
        return True


class IsAuthenticatedOrReadOnly(DjangoAuthorization):
    def read_list(self, object_list, bundle):
        return object_list

    def read_detail(self, object_list, bundle):
        return True


class PVInverter(models.Model):
    """
    Sandia model PV inverter parameters.
    """
    Sandia_ID = models.IntegerField()
    manufacturer = models.CharField('mfg.', max_length=100)
    name = models.CharField(max_length=100)
    source = models.CharField(max_length=10)
    vintage = models.DateField(default=date(1999, 1, 1))
    Paco = models.FloatField(default=-999)
    Vaco = models.FloatField(default=-999)
    Pdco = models.FloatField(default=-999)
    Vdco = models.FloatField(default=-999)
    Pso = models.FloatField(default=-999)
    C0 = models.FloatField(default=-999)
    C1 = models.FloatField(default=-999)
    C2 = models.FloatField(default=-999)
    C3 = models.FloatField(default=-999)
    Vdcmax = models.FloatField(default=-999)
    Idcmax = models.FloatField(default=-999)
    MPPT_low = models.FloatField(default=-999)
    MPPT_hi = models.FloatField(default=-999)
    Pnt = models.FloatField(default=-999)
    Tamb_low = models.FloatField(default=-999)
    Tamb_max = models.FloatField(default=-999)
    weight = models.FloatField(default=-999)
    numberMPPTChannels = models.IntegerField('number of MPPT channels',
                                             default=1)

    def full_name(self):
        return "%s %s (%d) - %d" % (self.manufacturer, self.name, self.Vaco,
                                    self.vintage.year)

    def __unicode__(self):
        return self.full_name()

    class Meta:
        verbose_name = "Inverter"
        unique_together = ('manufacturer', 'name', 'Vaco', 'vintage')

    @classmethod
    def upload(cls, filename, sheet=None, mapping=MAPPING):
        """
        Class method for creating and updating records from file.

        :param filename: Name of file.
        :param sheet: Name of sheet (None).
        :param mapping: Map of worksheet headers to model fields.
        :type mapping: dict
        """
        # load workbook
        wb = load_workbook(filename, use_iterators = True)
        # use first sheet by default
        if not sheet:
            sheet = wb.get_sheet_names()[0]
        # load worksheet
        ws = wb.get_sheet_by_name(sheet)
        rows = ws.iter_rows()  # row iterator
        # get headers
        headers = [row.value for row in rows.next() if row.value in mapping]
        # check headers match mapping
        if len(headers) < len(mapping):
            raise ValidationError('There are missing or incorrect headers.')
        # iterate over all rows
        for row in rows:
            # using Python-2.6 - doesn't have dictionary comprehension
            # zip will limit size to smaller list, ie: headers
            kwargs = dict([(mapping[h], r.value) for h, r in zip(headers, row)
                           if r.value])
            # convert vintage to date
            if 'vintage' in kwargs:
                kwargs['vintage'] = date(kwargs['vintage'], 1, 1)
            # ducktype temperature range for floats, pop otherwise
            should_be_floats = ('Tamb_low', 'Tamb_max', 'weight', 'Pnt')
            for val in should_be_floats:
                if val in kwargs:
                    try:
                        float(kwargs[val])
                    except ValueError:
                        kwargs.pop(val)
                        # TODO: put these in a "notes" catchall field
            try:
                # create new PVInverter record
                pvinv, created = cls.objects.get_or_create(**kwargs)
            except (IntegrityError, ValidationError) as err:
                # get key of existing record
                vintage = kwargs.get('vintage', date(1999, 1, 1))
                Vaco = kwargs.get('Vaco', -999)
                if cls.objects.filter(manufacturer=kwargs['manufacturer'],
                                      name=kwargs['name'], Vaco=Vaco,
                                      vintage=vintage).exists():
                    if Vaco != -999:
                        kwargs.pop('Vaco')
                    if vintage != date(1999, 1, 1):
                        kwargs.pop('vintage')
                    pvinv = cls.objects.get(
                        manufacturer=kwargs.pop('manufacturer'),
                        name=kwargs.pop('name'),Vaco=Vaco, vintage=vintage)
                    # update existing record
                    for k, v in kwargs.iteritems():
                        setattr(pvinv, k, v)
                    pvinv.save()
                    created = True
                else:
                    raise err  # raise caught exception
            if created:
                # TODO: use logger
                print 'created or updated %s' % pvinv
            else:
                print '%s already exists' % pvinv
            created = False


class PVInverterResource(ModelResource):
    class Meta:
        queryset = PVInverter.objects.all()
        filtering = {
            "manufacturer": (
                'iexact', 'istartswith', 'icontains', 'iregex', 'iendswith'
            ),
            "name": (
                'iexact', 'istartswith', 'icontains', 'iregex', 'iendswith'
            ),
            "Vaco": ('exact', 'lt', 'lte', 'gt', 'gte'),
            "vintage": ('year')
        }
        authorization = IsAuthenticatedOrReadOnly()
        authentication = ApiKeyAuthOrReadOnly()


class PVModule(models.Model):
    """
    Sandia model PV module parameters.
    """

    MATERIALS = [
        (1, '2-a-Si'),
        (2, '3-a-Si'),
        (3, 'CIS'),
        (4, 'CdTe'),
        (5, 'EFG'),
        (6, 'GaAs'),
        (7, 'HIT-Si'),
        (8, 'Si-Film'),
        (9, 'a-Si'),
        (10, 'c-Si'),
        (11, 'mc-Si'),
        (12, 'mono-Si')
    ]

    name = models.CharField(max_length=100)
    vintage = models.DateField(default=date(1999, 1, 1))
    vintage_estimated = models.BooleanField(default=False)
    area = models.FloatField(default=-999)
    material = models.IntegerField(choices=MATERIALS, default=10)
    cells_in_series = models.IntegerField(default=-999)
    parallel_strings = models.IntegerField(default=-999)
    fd = models.FloatField('diffuse fraction', default=-999)
    isc0 = models.FloatField(default=-999)
    voc0 = models.FloatField(default=-999)
    imp0 = models.FloatField(default=-999)
    vmp0 = models.FloatField(default=-999)
    ix0 = models.FloatField(default=-999)
    ixx0 = models.FloatField(default=-999)
    c0 = models.FloatField(default=-999)
    c1 = models.FloatField(default=-999)
    c2 = models.FloatField(default=-999)
    c3 = models.FloatField(default=-999)
    c4 = models.FloatField(default=-999)
    c5 = models.FloatField(default=-999)
    c6 = models.FloatField(default=-999)
    c7 = models.FloatField(default=-999)
    aisc = models.FloatField(default=-999)
    aimp = models.FloatField(default=-999)
    bvoc0 = models.FloatField(default=-999)
    mbvoc = models.FloatField(default=-999)
    bvmp0 = models.FloatField(default=-999)
    mbvmp = models.FloatField(default=-999)
    n = models.FloatField('ideality', default=-999)
    a0 = models.FloatField(default=-999)
    a1 = models.FloatField(default=-999)
    a2 = models.FloatField(default=-999)
    a3 = models.FloatField(default=-999)
    a4 = models.FloatField(default=-999)
    b0 = models.FloatField(default=-999)
    b1 = models.FloatField(default=-999)
    b2 = models.FloatField(default=-999)
    b3 = models.FloatField(default=-999)
    b4 = models.FloatField(default=-999)
    b5 = models.FloatField(default=-999)
    dt = models.FloatField(default=-999)
    a = models.FloatField('natural convection', default=-999)
    b = models.FloatField('forced convection', default=-999)
    notes = models.CharField(max_length=100)

    def nameplate(self):
        return self.imp0 * self.vmp0

    def fill_factor(self):
        return self.nameplate() / self.isc0 / self.voc0

    def module_eff(self):
        return self.nameplate() / self.area / 1000.0

    def __unicode__(self):
        return self.name

    class Meta:
        verbose_name = "Module"
        unique_together = ('name', 'vintage', 'vintage_estimated', 'notes')


class PVModuleResource(ModelResource):
    class Meta:
        queryset = PVModule.objects.all()
        filtering = {
            "name": (
                'iexact', 'istartswith', 'icontains', 'iregex', 'iendswith'
            ),
            "nameplate": ('exact', 'lt', 'lte', 'gt', 'gte'),
            "vintage": ('year')
        }
        authorization = IsAuthenticatedOrReadOnly()
        authentication = ApiKeyAuthOrReadOnly()
